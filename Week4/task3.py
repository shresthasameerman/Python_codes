import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

root = tk.Tk()
root.title("Data Store System")
root.geometry("700x600")
FILE_NAME = "Task3.xlsx"

def initialize_excel():
    if not os.path.exists(FILE_NAME):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "Age"])
        workbook.save(FILE_NAME)
        print("No file found, created a new Excel file.")
    else:
        print("File found.")

initialize_excel()

ui_frame = tk.Frame(root)
ui_frame.grid(row=1, column=0, sticky='nsew')

# Configure row and column weights for responsiveness
root.grid_rowconfigure(1, weight=1)  # Allow row 1 to expand
root.grid_columnconfigure(0, weight=1)  # Allow column 0 to expand

# Entry fields for Name and Age
tk.Label(ui_frame, text="Name").grid(row=0, column=0, sticky='ew')
entry_name = tk.Entry(ui_frame, width=40)  # Increased width for Name entry
entry_name.grid(row=0, column=1, sticky='ew')

tk.Label(ui_frame, text="Age").grid(row=1, column=0, sticky='ew')
enter_age = tk.Entry(ui_frame, width=40)  # Increased width for Age entry
enter_age.grid(row=1, column=1, sticky='ew')

# Create Treeview table
table = ttk.Treeview(ui_frame, columns=("Name", "Age"), show="headings")
table.heading("Name", text="Name ")
table.heading("Age", text="Age")
table.grid(row=3, column=0, columnspan=3, sticky='nsew')

# Configure Treeview to expand
ui_frame.grid_rowconfigure(3, weight=1)  # Allow row 3 to expand
ui_frame.grid_columnconfigure(0, weight=1)  # Allow column 0 to expand

# Load existing data into the Treeview
def load_data():
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip header
        table.insert("", "end", values=row)

load_data()  # Call load_data() after the table is defined

# Function to save data and insert it into the table
def save_data():
    name = entry_name.get()
    age = enter_age.get()

    # Insert the data into the Treeview table
    table.insert("", "end", values=(name, age))
    
    # Save the data to the Excel file
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    worksheet.append([name, age])  # Append the new data
    workbook.save(FILE_NAME)  # Save the workbook
    
    # Clear the entry fields after saving
    entry_name.delete(0, tk.END)
    enter_age.delete(0, tk.END)
    
    print("Your data has been saved.")

# Function to update selected entry
def update_data():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Update Error", "Please select an entry to update.")
        return

    name = entry_name.get()
    age = enter_age.get()

    if not name or not age:
        messagebox.showwarning("Input Error", "Please provide both Name and Age.")
        return

    # Update the Treeview
    table.item(selected_item, values=(name, age))

    # Update the Excel file
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active
    index = table.index(selected_item) + 2  # +2 to account for header and 0-indexing
    worksheet.cell(row=index, column=1, value=name)
    worksheet.cell(row=index, column=2, value=age)
    workbook.save(FILE_NAME)

    # Clear entry fields
    entry_name.delete(0, tk.END)
    enter_age.delete(0, tk.END)
    
    print("Your data has been updated.")

# Function to delete selected entry
def delete_data():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Delete Error", "Please select an entry to delete.")
        return

    # Get the values of the selected item to find its row in the Excel file
    item_values = table.item(selected_item, 'values')  # Retrieve the item values here
    name, age = item_values  # Assuming name and age are the first two columns

    # Remove the selected entry from the Treeview
    table.delete(selected_item)

    # Delete from Excel file
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook.active

    # Find the row to delete based on the name and age (assuming they are unique)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == age:
            worksheet.delete_rows(row[0].row)  # Delete the row
            break

    workbook.save(FILE_NAME)

    print("Your data has been deleted.")

# Button to save data
tk.Button(ui_frame, text="Save", command=save_data, width=10).grid(row=2, column=0, sticky='ew')  # Reduced button size

# Button to update data
tk.Button(ui_frame, text="Update", command=update_data, width=10).grid(row=2, column=1, sticky='ew')

# Button to delete data
tk.Button(ui_frame, text="Delete", command=delete_data, width=10).grid(row=2, column=2, sticky='ew')

# Bind double-click event to load selected entry into the entry fields for editing
def on_item_double_click(event):
    selected_item = table.selection()
    if selected_item:
        item_values = table.item(selected_item, 'values')
        entry_name.delete(0, tk.END)
        entry_name.insert(0, item_values[0])
        enter_age.delete(0, tk.END)
        enter_age.insert(0, item_values[1])

table.bind("<Double-1>", on_item_double_click)

root.mainloop()
